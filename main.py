import webview
import json
from pathlib import Path

DATA_FILE = Path('data.json')

def get_default_data():
    return {
        'settings': {
            'days': ['شنبه', 'یکشنبه', 'دوشنبه', 'سه‌شنبه', 'چهارشنبه'],
            'periodsCount': 4,
            'hoursPerPeriod': 2
        },
        'grades': ['دهم', 'یازدهم', 'دوازدهم'],
        'majors': ['ریاضی', 'تجربی', 'انسانی', 'کامپیوتر', 'حسابداری', 'عمومی'],
        'locations': [],
        'classes': [],
        'courses': [],
        'teachers': [],
        'allocations': [],
        'timetable': []
    }

class Api:
    def _read_data(self):
        if not DATA_FILE.exists():
            return get_default_data()
        try:
            with open(DATA_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                # Ensure all keys exist
                default = get_default_data()
                for k, v in default.items():
                    if k not in data:
                        data[k] = v
                return data
        except:
            return get_default_data()

    def _write_data(self, data):
        with open(DATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)

    def get_all_data(self):
        return self._read_data()

    def save_settings(self, settings_data):
        data = self._read_data()
        if 'settings' in settings_data:
            data['settings'] = settings_data['settings']
        if 'grades' in settings_data:
            data['grades'] = settings_data['grades']
        if 'majors' in settings_data:
            data['majors'] = settings_data['majors']
        self._write_data(data)

    def save_locations(self, locations):
        data = self._read_data()
        data['locations'] = locations
        self._write_data(data)

    def save_classes(self, classes):
        data = self._read_data()
        data['classes'] = classes
        self._write_data(data)

    def save_courses(self, courses):
        data = self._read_data()
        data['courses'] = courses
        self._write_data(data)

    def save_teachers(self, teachers):
        data = self._read_data()
        data['teachers'] = teachers
        self._write_data(data)

    def save_allocations(self, allocations):
        data = self._read_data()
        data['allocations'] = allocations
        self._write_data(data)

    def save_timetable(self, timetable):
        data = self._read_data()
        data['timetable'] = timetable
        self._write_data(data)
        
    def export_to_excel(self, data):
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            
            # Show save file dialog
            file_types = ('Excel Files (*.xlsx)', 'All files (*.*)')
            save_path = webview.windows[0].create_file_dialog(webview.SAVE_DIALOG, directory='', save_filename='SchoolMaster_Timetable.xlsx', file_types=file_types)
            
            if not save_path:
                return {'success': False, 'message': 'عملیات لغو شد.'}
            
            save_path = save_path[0] if isinstance(save_path, (list, tuple)) else save_path
            
            wb = openpyxl.Workbook()
            
            # Helper to style headers
            def style_header(cell):
                cell.font = Font(bold=True, name='Vazirmatn')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                thin = Side(border_style="thin", color="000000")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            def style_cell(cell, is_even=False, is_odd=False):
                cell.font = Font(name='Vazirmatn')
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                if is_even:
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # light green
                elif is_odd:
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # light yellow
                thin = Side(border_style="thin", color="000000")
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

            days = data['settings']['days']
            periods = data['settings']['periodsCount']
            
            def get_tt(day, period, filter_fn):
                tts = [t for t in data['timetable'] if t['day'] == day and t['period'] == period and filter_fn(t)]
                if not tts:
                    return ""
                
                texts = []
                is_even = any(t['weekType'] == 'even' for t in tts)
                is_odd = any(t['weekType'] == 'odd' for t in tts)
                
                for t in tts:
                    alloc = next((a for a in data['allocations'] if a['id'] == t['allocationId']), None)
                    if not alloc: continue
                    crs = next((c for c in data['courses'] if c['id'] == alloc['courseId']), None)
                    tch = next((tch for tch in data['teachers'] if tch['id'] == alloc['teacherId']), None)
                    cls = next((c for c in data['classes'] if c['id'] == alloc['classId']), None)
                    loc = next((l for l in data['locations'] if l['id'] == t.get('locationId')), None)
                    
                    w_str = '(زوج)' if t['weekType'] == 'even' else '(فرد)' if t['weekType'] == 'odd' else ''
                    loc_str = f" [{loc['name']}]" if loc else ""
                    
                    texts.append(f"{crs['name']} {w_str}\n{tch['name']} - {cls['name']}{loc_str}")
                
                return ("\n\n".join(texts), is_even, is_odd)

            # Sheet 1: Master
            ws_master = wb.active
            ws_master.title = "برنامه کلی"
            ws_master.sheet_view.rightToLeft = True
            
            ws_master.cell(row=1, column=1, value="کلاس / روز")
            style_header(ws_master.cell(row=1, column=1))
            col_idx = 2
            for d in days:
                cell = ws_master.cell(row=1, column=col_idx, value=d)
                ws_master.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+periods-1)
                style_header(cell)
                for p in range(1, periods + 1):
                    c = ws_master.cell(row=2, column=col_idx + p - 1, value=f"زنگ {p}")
                    style_header(c)
                col_idx += periods
            
            row_idx = 3
            for cls in data['classes']:
                ws_master.cell(row=row_idx, column=1, value=f"{cls['name']}\n{cls['major']}")
                style_cell(ws_master.cell(row=row_idx, column=1))
                c_idx = 2
                for d in days:
                    for p in range(1, periods + 1):
                        def f_cls(t, c_id=cls['id']):
                            a = next((a for a in data['allocations'] if a['id'] == t['allocationId']), None)
                            return a and a['classId'] == c_id
                        val, ev, od = get_tt(d, p, f_cls)
                        cell = ws_master.cell(row=row_idx, column=c_idx, value=val if val else "-")
                        style_cell(cell, is_even=ev, is_odd=od)
                        c_idx += 1
                row_idx += 1

            ws_master.column_dimensions['A'].width = 15
            for col in range(2, 2 + len(days) * periods):
                ws_master.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

            # Sheet 2: Classes
            ws_classes = wb.create_sheet(title="برنامه کلاس‌ها")
            ws_classes.sheet_view.rightToLeft = True
            
            r_idx = 1
            for cls in data['classes']:
                ws_classes.cell(row=r_idx, column=1, value=f"کلاس: {cls['name']} ({cls['major']})")
                ws_classes.cell(row=r_idx, column=1).font = Font(bold=True, size=14, name='Vazirmatn')
                r_idx += 1
                
                ws_classes.cell(row=r_idx, column=1, value="روز")
                style_header(ws_classes.cell(row=r_idx, column=1))
                for p in range(1, periods + 1):
                    c = ws_classes.cell(row=r_idx, column=p + 1, value=f"زنگ {p}")
                    style_header(c)
                r_idx += 1
                
                for d in days:
                    ws_classes.cell(row=r_idx, column=1, value=d)
                    style_header(ws_classes.cell(row=r_idx, column=1))
                    for p in range(1, periods + 1):
                        def f_cls(t, c_id=cls['id']):
                            a = next((a for a in data['allocations'] if a['id'] == t['allocationId']), None)
                            return a and a['classId'] == c_id
                        val, ev, od = get_tt(d, p, f_cls)
                        cell = ws_classes.cell(row=r_idx, column=p + 1, value=val if val else "-")
                        style_cell(cell, is_even=ev, is_odd=od)
                    r_idx += 1
                r_idx += 2 

            ws_classes.column_dimensions['A'].width = 15
            for col in range(2, periods + 2):
                ws_classes.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

            # Sheet 3: Teachers
            ws_teachers = wb.create_sheet(title="برنامه دبیران")
            ws_teachers.sheet_view.rightToLeft = True
            
            r_idx = 1
            for tch in data['teachers']:
                ws_teachers.cell(row=r_idx, column=1, value=f"دبیر: {tch['name']}")
                ws_teachers.cell(row=r_idx, column=1).font = Font(bold=True, size=14, name='Vazirmatn')
                r_idx += 1
                
                ws_teachers.cell(row=r_idx, column=1, value="روز")
                style_header(ws_teachers.cell(row=r_idx, column=1))
                for p in range(1, periods + 1):
                    c = ws_teachers.cell(row=r_idx, column=p + 1, value=f"زنگ {p}")
                    style_header(c)
                r_idx += 1
                
                for d in days:
                    ws_teachers.cell(row=r_idx, column=1, value=d)
                    style_header(ws_teachers.cell(row=r_idx, column=1))
                    for p in range(1, periods + 1):
                        def f_tch(t, t_id=tch['id']):
                            a = next((a for a in data['allocations'] if a['id'] == t['allocationId']), None)
                            return a and a['teacherId'] == t_id
                        val, ev, od = get_tt(d, p, f_tch)
                        cell = ws_teachers.cell(row=r_idx, column=p + 1, value=val if val else "-")
                        style_cell(cell, is_even=ev, is_odd=od)
                    r_idx += 1
                r_idx += 2 

            ws_teachers.column_dimensions['A'].width = 15
            for col in range(2, periods + 2):
                ws_teachers.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

            wb.save(save_path)
            return {'success': True, 'message': 'فایل اکسل با موفقیت ذخیره شد.'}
            
        except Exception as e:
            return {'success': False, 'message': f'خطا در ذخیره فایل: {str(e)}'}

    def clear_all_data(self):
        self._write_data(get_default_data())


if __name__ == '__main__':
    api = Api()
    
    base_dir = Path(__file__).parent
    html_file = base_dir / 'index.html'
    
    if not html_file.exists():
        with open(html_file, 'w', encoding='utf-8') as f:
            f.write("<h1>Loading...</h1>")
    
    window = webview.create_window(
        title='سیستم برنامه‌ریزی و مدیریت مدرسه',
        url=str(html_file),
        js_api=api,
        width=1200,
        height=800,
        min_size=(800, 600)
    )
    
    webview.start()
